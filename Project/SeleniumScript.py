import openpyxl
import google.generativeai as genai
from openpyxl.styles import Font, Alignment

# Configure Gemini
GOOGLE_API_KEY = "AIzaSyAaSko_ML10ndIOzoK7h3JDIs_nlEBTh6w"  # Replace with your API key
genai.configure(api_key=GOOGLE_API_KEY)
model = genai.GenerativeModel('gemini-2.0-flash')

# Load test cases from previous task
test_cases_wb = openpyxl.load_workbook('test_cases.xlsx')
test_cases_ws = test_cases_wb.active

# Prepare new workbook for scripts
scripts_wb = openpyxl.Workbook()
scripts_ws = scripts_wb.active
scripts_ws.title = "Test Scripts"
scripts_ws.append(["Test Case ID", "Python Selenium Code"])

# Configure cell styling
wrap_alignment = Alignment(wrap_text=True)
font = Font(name='Courier New', size=10)

def generate_selenium_code(test_case):
    prompt = f"""
    Generate a Python Selenium script for the following test case. Follow these requirements:
    1. Use Chrome WebDriver
    2. Assume the base URL is https://demoblaze.com/
    3. Use explicit waits with WebDriverWait
    4. Include proper setup and teardown
    5. Use ID locators when available, otherwise use CSS selectors
    6. Include assertions for expected results
    
    Test Case:
    ID: {test_case['id']}
    Scenario: {test_case['scenario']}
    Steps: {test_case['steps']}
    Expected Result: {test_case['expected']}
    
    Return ONLY the complete Python code without any explanations.
    """
    
    response = model.generate_content(prompt)
    return response.text

# Process each test case
for row in test_cases_ws.iter_rows(min_row=2, values_only=True):
    test_case = {
        'id': row[0],
        'scenario': row[1],
        'steps': row[2],
        'expected': row[3]
    }
    
    print(f"Generating script for {test_case['id']}...")
    try:
        selenium_code = generate_selenium_code(test_case)
        # Clean up code formatting
        selenium_code = "\n".join([line.rstrip() for line in selenium_code.split('\n')])
        
        # Add to worksheet
        scripts_ws.append([test_case['id'], selenium_code])
        
        # Apply formatting to code cell
        code_cell = scripts_ws.cell(row=scripts_ws.max_row, column=2)
        code_cell.alignment = wrap_alignment
        code_cell.font = font
    
    except Exception as e:
        print(f"Error generating script for {test_case['id']}: {str(e)}")
        scripts_ws.append([test_case['id'], f"Error: {str(e)}"])

# Adjust column widths
scripts_ws.column_dimensions['A'].width = 15
scripts_ws.column_dimensions['B'].width = 120

# Save the workbook
scripts_wb.save("test_scripts.xlsx")
print("Selenium scripts generated successfully in test_scripts.xlsx")